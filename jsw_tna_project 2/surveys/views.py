import io
import json
from datetime import date
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone

from .models import TNAFHQResponse, TNASelfAssessmentResponse
from .competencies import (
    FUNCTIONS, ROLE_CHOICES, FUNCTION_CHOICES,
    PROFICIENCY_SCALE, IMPORTANCE_SCALE, FHQ_OPEN_QUESTIONS,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def get_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    return xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR')


def admin_required(fn):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('admin_logged_in'):
            return redirect('/admin-login/')
        return fn(request, *args, **kwargs)
    return wrapper


# ── welcome ──────────────────────────────────────────────────────────────────

def welcome(request):
    if request.method == 'POST':
        name     = request.POST.get('name', '').strip()
        emp_code = request.POST.get('emp_code', '').strip()
        role     = request.POST.get('role', '').strip()
        function = request.POST.get('function', '').strip()

        if not all([name, emp_code, role, function]):
            return render(request, 'surveys/welcome.html', {
                'error': 'All fields are required.',
                'functions': FUNCTION_CHOICES, 'roles': ROLE_CHOICES,
                'post': request.POST,
            })

        request.session['user_name']     = name
        request.session['user_emp_code'] = emp_code
        request.session['user_role']     = role
        request.session['user_function'] = function

        if role == 'FH':
            return redirect('/survey/fhq/')
        elif role == 'GET':
            return redirect('/survey/get/self-assessment/')
        else:  # MT
            return redirect('/survey/mt/self-assessment/')

    return render(request, 'surveys/welcome.html', {
        'functions': FUNCTION_CHOICES,
        'roles': ROLE_CHOICES,
    })


# ── FHQ ──────────────────────────────────────────────────────────────────────

def fhq(request):
    name     = request.session.get('user_name', '')
    emp_code = request.session.get('user_emp_code', '')
    function = request.session.get('user_function', '')

    if not function or function not in FUNCTIONS:
        return redirect('/')

    fn_data = FUNCTIONS[function]
    comps   = fn_data['comps']

    if request.method == 'POST':
        ratings = {}
        for c in comps:
            imp  = request.POST.get(f'imp_{c["code"]}')
            dget = request.POST.get(f'dget_{c["code"]}')
            dmt  = request.POST.get(f'dmt_{c["code"]}')
            if imp and dget and dmt:
                ratings[c['code']] = {
                    'name': c['name'],
                    'importance': int(imp),
                    'desired_get': int(dget),
                    'desired_mt': int(dmt),
                }

        try:
            TNAFHQResponse.objects.create(
                employee_name=request.POST.get('employee_name', name).strip(),
                employee_code=request.POST.get('employee_code', emp_code).strip(),
                role='FH',
                function=function,
                designation=request.POST.get('designation', '').strip(),
                department=request.POST.get('department', '').strip(),
                date=request.POST.get('date') or date.today(),
                competency_ratings=ratings,
                q_critical=request.POST.get('q_critical', '').strip(),
                q_gap_risk=request.POST.get('q_gap_risk', '').strip(),
                q_ojt=request.POST.get('q_ojt', '').strip(),
                q_alp=request.POST.get('q_alp', '').strip(),
                ip_address=get_ip(request),
            )
            return redirect('/survey/success/?type=fhq')
        except Exception as e:
            return render(request, 'surveys/fhq.html', {
                'error': str(e), 'fn_data': fn_data, 'comps': comps,
                'open_qs': FHQ_OPEN_QUESTIONS, 'prof_scale': PROFICIENCY_SCALE,
                'imp_scale': IMPORTANCE_SCALE, 'name': name, 'emp_code': emp_code,
                'function_key': function,
            })

    return render(request, 'surveys/fhq.html', {
        'fn_data': fn_data, 'comps': comps, 'open_qs': FHQ_OPEN_QUESTIONS,
        'prof_scale': PROFICIENCY_SCALE, 'imp_scale': IMPORTANCE_SCALE,
        'name': name, 'emp_code': emp_code, 'function_key': function,
    })


# ── Self-Assessment ───────────────────────────────────────────────────────────

def self_assessment(request, role):
    role = role.upper()
    name     = request.session.get('user_name', '')
    emp_code = request.session.get('user_emp_code', '')
    function = request.session.get('user_function', '')

    if not function or function not in FUNCTIONS:
        return redirect('/')

    fn_data = FUNCTIONS[function]
    comps   = fn_data['comps']

    if request.method == 'POST':
        ratings = {}
        for c in comps:
            val = request.POST.get(f'level_{c["code"]}')
            if val:
                ratings[c['code']] = {
                    'name': c['name'],
                    'current_level': int(val),
                }
        try:
            TNASelfAssessmentResponse.objects.create(
                employee_name=request.POST.get('employee_name', name).strip(),
                employee_code=request.POST.get('employee_code', emp_code).strip(),
                role=role,
                function=function,
                date=request.POST.get('date') or date.today(),
                competency_ratings=ratings,
                ip_address=get_ip(request),
            )
            return redirect(f'/survey/success/?type=sa_{role.lower()}')
        except Exception as e:
            return render(request, 'surveys/self_assessment.html', {
                'error': str(e), 'role': role, 'fn_data': fn_data, 'comps': comps,
                'prof_scale': PROFICIENCY_SCALE, 'name': name, 'emp_code': emp_code,
                'function_key': function,
            })

    return render(request, 'surveys/self_assessment.html', {
        'role': role, 'fn_data': fn_data, 'comps': comps,
        'prof_scale': PROFICIENCY_SCALE, 'name': name, 'emp_code': emp_code,
        'function_key': function,
    })


# ── Success ───────────────────────────────────────────────────────────────────

def success(request):
    stype = request.GET.get('type', 'survey')
    labels = {
        'fhq':    ('Functional Head Questionnaire', '📋', '#1A2D5A'),
        'sa_get': ('GET Self-Assessment', '👤', '#0F766E'),
        'sa_mt':  ('MT Self-Assessment',  '👤', '#7C3AED'),
    }
    title, icon, color = labels.get(stype, ('Survey', '✅', '#28a745'))
    return render(request, 'surveys/success.html', {
        'title': title, 'icon': icon, 'color': color,
    })


# ── Admin Login / Logout ──────────────────────────────────────────────────────

def admin_login(request):
    if request.session.get('admin_logged_in'):
        return redirect('/admin-dashboard/')
    error = None
    if request.method == 'POST':
        u = request.POST.get('username', '').strip()
        p = request.POST.get('password', '').strip()
        if u == settings.ADMIN_USERNAME and p == settings.ADMIN_PASSWORD:
            request.session['admin_logged_in'] = True
            return redirect('/admin-dashboard/')
        error = 'Invalid credentials.'
    return render(request, 'surveys/admin_login.html', {'error': error})


def admin_logout(request):
    request.session.flush()
    return redirect('/admin-login/')


# ── Admin Dashboard ───────────────────────────────────────────────────────────

@admin_required
def admin_dashboard(request):
    fhq_total = TNAFHQResponse.objects.count()
    sa_total  = TNASelfAssessmentResponse.objects.count()

    # Breakdown by function
    fn_stats = []
    for fk, fv in FUNCTIONS.items():
        fhq_count = TNAFHQResponse.objects.filter(function=fk).count()
        get_count = TNASelfAssessmentResponse.objects.filter(function=fk, role='GET').count()
        mt_count  = TNASelfAssessmentResponse.objects.filter(function=fk, role='MT').count()
        fn_stats.append({
            'key': fk, 'label': fv['label'], 'icon': fv['icon'],
            'fhq': fhq_count, 'get': get_count, 'mt': mt_count,
        })

    recent_fhq = TNAFHQResponse.objects.order_by('-submitted_at')[:8]
    recent_sa  = TNASelfAssessmentResponse.objects.order_by('-submitted_at')[:8]

    return render(request, 'surveys/admin_dashboard.html', {
        'fhq_total': fhq_total, 'sa_total': sa_total,
        'total': fhq_total + sa_total,
        'fn_stats': fn_stats,
        'recent_fhq': recent_fhq, 'recent_sa': recent_sa,
        'functions': FUNCTIONS,
    })


# ── Admin – Responses List ────────────────────────────────────────────────────

@admin_required
def admin_responses(request, survey_type):
    fn_filter   = request.GET.get('function', '')
    role_filter = request.GET.get('role', '')

    if survey_type == 'fhq':
        qs = TNAFHQResponse.objects.all()
        if fn_filter:
            qs = qs.filter(function=fn_filter)
        title = 'Functional Head Questionnaire Responses'
    elif survey_type == 'self-assessment':
        qs = TNASelfAssessmentResponse.objects.all()
        if fn_filter:
            qs = qs.filter(function=fn_filter)
        if role_filter in ('GET', 'MT'):
            qs = qs.filter(role=role_filter)
        title = 'Self-Assessment Responses'
    else:
        return redirect('/admin-dashboard/')

    return render(request, 'surveys/admin_responses.html', {
        'responses': qs, 'title': title,
        'survey_type': survey_type,
        'fn_filter': fn_filter, 'role_filter': role_filter,
        'functions': FUNCTION_CHOICES,
    })


# ── Gap Analysis ──────────────────────────────────────────────────────────────

@admin_required
def gap_analysis(request):
    fn_key = request.GET.get('function', list(FUNCTIONS.keys())[0])
    if fn_key not in FUNCTIONS:
        fn_key = list(FUNCTIONS.keys())[0]

    fn_data = FUNCTIONS[fn_key]
    comps   = fn_data['comps']

    fhq_qs = TNAFHQResponse.objects.filter(function=fn_key)
    get_qs  = TNASelfAssessmentResponse.objects.filter(function=fn_key, role='GET')
    mt_qs   = TNASelfAssessmentResponse.objects.filter(function=fn_key, role='MT')

    analysis = []
    for c in comps:
        code = c['code']

        # FHQ desired levels
        imp_vals, dget_vals, dmt_vals = [], [], []
        for r in fhq_qs:
            d = r.competency_ratings.get(code, {})
            if d.get('importance'): imp_vals.append(d['importance'])
            if d.get('desired_get'): dget_vals.append(d['desired_get'])
            if d.get('desired_mt'):  dmt_vals.append(d['desired_mt'])

        avg_imp  = round(sum(imp_vals)  / len(imp_vals),  1) if imp_vals  else None
        avg_dget = round(sum(dget_vals) / len(dget_vals), 1) if dget_vals else None
        avg_dmt  = round(sum(dmt_vals)  / len(dmt_vals),  1) if dmt_vals  else None

        # Self-assessment current levels
        get_vals, mt_vals = [], []
        for r in get_qs:
            v = r.competency_ratings.get(code, {}).get('current_level')
            if v: get_vals.append(v)
        for r in mt_qs:
            v = r.competency_ratings.get(code, {}).get('current_level')
            if v: mt_vals.append(v)

        avg_get = round(sum(get_vals) / len(get_vals), 1) if get_vals else None
        avg_mt  = round(sum(mt_vals)  / len(mt_vals),  1) if mt_vals  else None

        gap_get = round(avg_dget - avg_get, 1) if (avg_dget and avg_get) else None
        gap_mt  = round(avg_dmt  - avg_mt,  1) if (avg_dmt  and avg_mt)  else None

        analysis.append({
            'code': code, 'name': c['name'], 'category': c['category'],
            'ev_critical': c['ev_critical'],
            'avg_imp': avg_imp,
            'avg_dget': avg_dget, 'avg_get': avg_get, 'gap_get': gap_get,
            'avg_dmt': avg_dmt,  'avg_mt': avg_mt,   'gap_mt': gap_mt,
            'fhq_count': fhq_qs.count(),
            'get_count': len(get_vals), 'mt_count': len(mt_vals),
        })

    # Summary
    gaps_get = [a['gap_get'] for a in analysis if a['gap_get'] is not None]
    gaps_mt  = [a['gap_mt']  for a in analysis if a['gap_mt']  is not None]
    avg_gap_get = round(sum(gaps_get) / len(gaps_get), 1) if gaps_get else None
    avg_gap_mt  = round(sum(gaps_mt)  / len(gaps_mt),  1) if gaps_mt  else None

    return render(request, 'surveys/gap_analysis.html', {
        'fn_key': fn_key, 'fn_data': fn_data,
        'analysis': analysis,
        'avg_gap_get': avg_gap_get, 'avg_gap_mt': avg_gap_mt,
        'fhq_count': fhq_qs.count(),
        'get_count': get_qs.count(), 'mt_count': mt_qs.count(),
        'functions': FUNCTION_CHOICES,
    })


# ── Excel Export ──────────────────────────────────────────────────────────────

@admin_required
def export_excel(request, survey_type):
    fn_filter   = request.GET.get('function', '')
    role_filter = request.GET.get('role', '')

    wb = openpyxl.Workbook()

    H_FONT  = Font(bold=True, color='FFFFFF', size=10)
    H_FILL  = PatternFill('solid', fgColor='1A2D5A')
    SH_FILL = PatternFill('solid', fgColor='E8F0FE')
    CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    BORDER  = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'),  bottom=Side(style='thin'))

    def hdr(ws, row, n):
        for col in range(1, n + 1):
            c = ws.cell(row=row, column=col)
            c.font = H_FONT; c.fill = H_FILL
            c.alignment = CENTER; c.border = BORDER

    def stylerow(ws, row, n, alt=False):
        for col in range(1, n + 1):
            c = ws.cell(row=row, column=col)
            if alt: c.fill = SH_FILL
            c.alignment = LEFT; c.border = BORDER

    if survey_type == 'fhq':
        qs = TNAFHQResponse.objects.all()
        if fn_filter: qs = qs.filter(function=fn_filter)

        ws = wb.active; ws.title = 'FHQ Responses'
        fn_keys = list(FUNCTIONS.keys())
        # Dynamic competency columns
        sample_fn = fn_filter if fn_filter in FUNCTIONS else fn_keys[0]
        comp_codes = [c['code'] for c in FUNCTIONS[sample_fn]['comps']]

        base = ['#', 'Emp Name', 'Emp Code', 'Function', 'Designation', 'Dept', 'Date']
        imp_hdrs  = [f'{code}\nImportance' for code in comp_codes]
        dget_hdrs = [f'{code}\nDesired (GET)' for code in comp_codes]
        dmt_hdrs  = [f'{code}\nDesired (MT)' for code in comp_codes]
        open_hdrs = ['Q1: Most Critical Competency', 'Q2: Biggest Gap Risk',
                     'Q3: Best OJT Experience', 'Q4: ALP Theme Suggestion']
        headers = base + imp_hdrs + dget_hdrs + dmt_hdrs + open_hdrs + ['Submitted At']
        ws.append(headers); hdr(ws, 1, len(headers))

        for i, r in enumerate(qs, 1):
            row_data = [i, r.employee_name, r.employee_code, r.get_function_display(),
                        r.designation, r.department, str(r.date)]
            # Use function-specific comps if filtered
            fn_comps = FUNCTIONS.get(r.function, FUNCTIONS[sample_fn])['comps']
            fn_codes = [c['code'] for c in fn_comps]
            row_data += [r.competency_ratings.get(c, {}).get('importance', '') for c in comp_codes]
            row_data += [r.competency_ratings.get(c, {}).get('desired_get', '') for c in comp_codes]
            row_data += [r.competency_ratings.get(c, {}).get('desired_mt', '') for c in comp_codes]
            row_data += [r.q_critical, r.q_gap_risk, r.q_ojt, r.q_alp]
            row_data.append(r.submitted_at.strftime('%Y-%m-%d %H:%M'))
            ws.append(row_data); stylerow(ws, i + 1, len(headers), i % 2 == 0)

    elif survey_type == 'self-assessment':
        qs = TNASelfAssessmentResponse.objects.all()
        if fn_filter:   qs = qs.filter(function=fn_filter)
        if role_filter in ('GET', 'MT'): qs = qs.filter(role=role_filter)

        ws = wb.active; ws.title = 'Self-Assessment Responses'
        sample_fn = fn_filter if fn_filter in FUNCTIONS else list(FUNCTIONS.keys())[0]
        comp_codes = [c['code'] for c in FUNCTIONS[sample_fn]['comps']]

        base    = ['#', 'Emp Name', 'Emp Code', 'Role', 'Function', 'Date']
        hdrs_cl = [f'{code}\nCurrent Level' for code in comp_codes]
        headers = base + hdrs_cl + ['Avg Score', 'Submitted At']
        ws.append(headers); hdr(ws, 1, len(headers))

        for i, r in enumerate(qs, 1):
            row_data = [i, r.employee_name, r.employee_code, r.role,
                        r.get_function_display(), str(r.date)]
            row_data += [r.competency_ratings.get(c, {}).get('current_level', '') for c in comp_codes]
            row_data += [r.avg_rating(), r.submitted_at.strftime('%Y-%m-%d %H:%M')]
            ws.append(row_data); stylerow(ws, i + 1, len(headers), i % 2 == 0)

    elif survey_type == 'gap-analysis':
        fn_key  = fn_filter if fn_filter in FUNCTIONS else list(FUNCTIONS.keys())[0]
        fn_data = FUNCTIONS[fn_key]
        comps   = fn_data['comps']
        fhq_qs  = TNAFHQResponse.objects.filter(function=fn_key)
        get_qs  = TNASelfAssessmentResponse.objects.filter(function=fn_key, role='GET')
        mt_qs   = TNASelfAssessmentResponse.objects.filter(function=fn_key, role='MT')

        ws = wb.active; ws.title = f'Gap Analysis — {fn_data["label"]}'
        headers = ['#', 'Code', 'Competency', 'Category', 'EV Critical',
                   'Avg Importance (FH)', 'Desired Level GET (FH)', 'Avg Current (GET)', 'GAP GET',
                   'Desired Level MT (FH)', 'Avg Current (MT)', 'GAP MT']
        ws.append(headers); hdr(ws, 1, len(headers))

        for i, c in enumerate(comps, 1):
            code = c['code']
            imp_v, dg_v, dm_v, cg_v, cm_v = [], [], [], [], []
            for r in fhq_qs:
                d = r.competency_ratings.get(code, {})
                if d.get('importance'): imp_v.append(d['importance'])
                if d.get('desired_get'): dg_v.append(d['desired_get'])
                if d.get('desired_mt'):  dm_v.append(d['desired_mt'])
            for r in get_qs:
                v = r.competency_ratings.get(code, {}).get('current_level')
                if v: cg_v.append(v)
            for r in mt_qs:
                v = r.competency_ratings.get(code, {}).get('current_level')
                if v: cm_v.append(v)

            avg_imp = round(sum(imp_v)/len(imp_v), 1) if imp_v else ''
            avg_dg  = round(sum(dg_v)/len(dg_v), 1) if dg_v else ''
            avg_dm  = round(sum(dm_v)/len(dm_v), 1) if dm_v else ''
            avg_cg  = round(sum(cg_v)/len(cg_v), 1) if cg_v else ''
            avg_cm  = round(sum(cm_v)/len(cm_v), 1) if cm_v else ''
            gap_g   = round(avg_dg - avg_cg, 1) if (avg_dg and avg_cg) else ''
            gap_m   = round(avg_dm - avg_cm, 1) if (avg_dm and avg_cm) else ''

            row_data = [i, code, c['name'], c['category'], 'Yes' if c['ev_critical'] else 'No',
                        avg_imp, avg_dg, avg_cg, gap_g, avg_dm, avg_cm, gap_m]
            ws.append(row_data); stylerow(ws, i + 1, len(headers), i % 2 == 0)

            # Color gap cells: red if gap > 1, yellow if 0 < gap ≤ 1, green if ≤ 0
            for col_idx, gap_val in [(9, gap_g), (12, gap_m)]:
                if isinstance(gap_val, (int, float)):
                    cell = ws.cell(row=i + 1, column=col_idx)
                    if gap_val > 1:
                        cell.fill = PatternFill('solid', fgColor='FFCCCC')
                    elif gap_val > 0:
                        cell.fill = PatternFill('solid', fgColor='FFF3CD')
                    else:
                        cell.fill = PatternFill('solid', fgColor='D4EDDA')

    for sheet in wb.worksheets:
        sheet.freeze_panes = 'A2'
        sheet.column_dimensions['B'].width = 14
        sheet.column_dimensions['C'].width = 36

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    from datetime import datetime
    fn_label = FUNCTIONS.get(fn_filter, {}).get('short', fn_filter or 'ALL')
    filename = f"JSW_TNA_{survey_type}_{fn_label}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
